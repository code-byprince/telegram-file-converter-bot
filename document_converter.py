import csv
import pdfplumber
from docx import Document
from pdf2docx import Converter
from PyPDF2 import PdfReader, PdfWriter
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet


def pdf_to_text(pdf_path: str, output_path: str):
    """PDF se text nikaal ke .txt file banata hai."""
    text_parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text_parts.append(page.extract_text() or "")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n\n".join(text_parts))
    return output_path


def text_to_pdf(text_path: str, output_path: str):
    """.txt file ko PDF me convert karta hai."""
    with open(text_path, "r", encoding="utf-8", errors="ignore") as f:
        content = f.read()

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                             leftMargin=2 * cm, rightMargin=2 * cm,
                             topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    story = []
    for line in content.split("\n"):
        safe_line = line.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        story.append(Paragraph(safe_line if safe_line.strip() else "&nbsp;", styles["Normal"]))
    doc.build(story)
    return output_path


def pdf_to_word(pdf_path: str, output_path: str):
    """PDF ko .docx me convert karta hai (layout ke saath, pdf2docx library)."""
    cv = Converter(pdf_path)
    cv.convert(output_path)
    cv.close()
    return output_path


def word_to_pdf(docx_path: str, output_path: str):
    """
    .docx ko PDF me convert karta hai (lightweight method).
    Note: Yeh basic text/paragraph formatting preserve karta hai.
    Complex tables/images wale docx ke liye kabhi formatting simple ho sakti hai
    (free hosting pe LibreOffice jaisa heavy tool avoid karne ke liye).
    """
    document = Document(docx_path)
    doc = SimpleDocTemplate(output_path, pagesize=A4,
                             leftMargin=2 * cm, rightMargin=2 * cm,
                             topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    story = []
    for para in document.paragraphs:
        text = para.text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        style_name = "Heading1" if para.style.name.startswith("Heading") else "Normal"
        story.append(Paragraph(text if text.strip() else "&nbsp;", styles.get(style_name, styles["Normal"])))
    doc.build(story)
    return output_path


def merge_pdfs(pdf_paths: list, output_path: str):
    """Multiple PDFs ko order me jodh ke ek PDF banata hai."""
    writer = PdfWriter()
    for path in pdf_paths:
        reader = PdfReader(path)
        for page in reader.pages:
            writer.add_page(page)
    with open(output_path, "wb") as f:
        writer.write(f)
    return output_path


def split_pdf(pdf_path: str, output_path: str, start_page: int, end_page: int):
    """1-indexed page range (inclusive) nikal ke naya PDF banata hai."""
    reader = PdfReader(pdf_path)
    writer = PdfWriter()
    total_pages = len(reader.pages)
    start = max(1, start_page)
    end = min(total_pages, end_page)
    if start > total_pages:
        raise ValueError(f"PDF me sirf {total_pages} pages hain, start page {start} nahi mila.")
    for i in range(start - 1, end):
        writer.add_page(reader.pages[i])
    with open(output_path, "wb") as f:
        writer.write(f)
    return output_path


def add_pdf_password(pdf_path: str, output_path: str, password: str):
    """PDF pe password laga deta hai."""
    reader = PdfReader(pdf_path)
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
    writer.encrypt(password)
    with open(output_path, "wb") as f:
        writer.write(f)
    return output_path


def remove_pdf_password(pdf_path: str, output_path: str, password: str):
    """PDF se password hata deta hai (sahi password chahiye)."""
    reader = PdfReader(pdf_path, password=password)
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
    with open(output_path, "wb") as f:
        writer.write(f)
    return output_path


def excel_to_csv(xlsx_path: str, output_path: str):
    """Excel (.xlsx) ki pehli sheet ko .csv me convert karta hai."""
    wb = load_workbook(xlsx_path, data_only=True)
    sheet = wb.active
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if cell is None else cell for cell in row])
    return output_path


def csv_to_excel(csv_path: str, output_path: str):
    """.csv file ko Excel (.xlsx) me convert karta hai."""
    wb = Workbook()
    sheet = wb.active
    with open(csv_path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        for row in reader:
            sheet.append(row)
    wb.save(output_path)
    return output_path


def get_page_count(pdf_path: str) -> int:
    """PDF me kitne pages hain, batata hai."""
    reader = PdfReader(pdf_path)
    return len(reader.pages)
